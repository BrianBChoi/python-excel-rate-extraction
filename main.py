import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.worksheet.copier import WorksheetCopy
from copy import copy

from rate_data import RateTable, Origin, Destination
import styles

rate_tables = []
container_types = {
    "20'": 49,
    "STD": 55,
    "HQ": 60,
    "45'": 66,
    "40'HRD": 72
}
keep_merged = ['C3:D4','H2:N2','H3:N3','H4:N4','H5:N5','O2:V3','O4:V5']
columns = {
    'Group': 1,
    'Origin': 2,
    'Destination': 3,
    'Via': 4,
    'Term': 5,
    'EQ': 6,
    'O/F': 7,
    'Sub Total': 18,
    'OAC': 19,
    'IHC': 20,
    'NAC': 22
}


def main():
    read_file('KR2212095AMD066.xls')
    write_file(r'output\template.xlsx', r'output\test.xlsx')

def read_file(filename):
    data = pd.read_excel(filename)

    for row in range(data.shape[0]):
        value = str(data.iloc[row, 1])
        if 'RATE FOR BULLET NO.' in value:
            value = value[11:]
            rt = RateTable(name=value)
            rate_tables.append(rt)
            print(rt.name)
        elif value == 'COMMODITY':
            do_while = True
            i = 0
            while do_while:
                code = str(data.iloc[row + i, 18])
                if code != 'nan':
                    rate_tables[-1].add_commodity_code(code)
                i += 1
                if(str(data.iloc[row + i, 1]) != 'nan'):
                    do_while = False
            print("COMMODITY: {}".format(rate_tables[-1].commodity_codes))
        elif value == 'ORIGIN':
            name = str(data.iloc[row, 18])
            origin = Origin(name)
            rate_tables[-1].add_origin(origin)
            print("ORIGIN: {}".format(rate_tables[-1].origins[-1].name))
        elif value == 'Destination':
            do_while = True
            i = 1
            while do_while:
                name = str(data.iloc[row + i, 1])
                term = str(data.iloc[row + i, 22])
                via = str(data.iloc[row + i, 27])
                destination = Destination(name, term, via)
                print("name: {}, term: {}, via: {}".format(name, term, via))
                rate_tables[-1].origins[-1].add_destination(destination)
                for type, col in container_types.items():
                    rate = str(data.iloc[row + i, col])
                    match = re.search(r'(\d+(,\d+)*)', rate)
                    if match:
                        number = int(match.group().replace(",",""))
                        destination.add_rate(type, number)
                print(rate_tables[-1].origins[-1].destinations[-1].rates)
                i += 1
                if(str(data.iloc[row + i, 1]) == 'nan'):
                    do_while = False


def write_file(input_file, output_file):
    print("Opening file")
    wb = load_workbook(input_file)

    print("Formatting new sheet")
    old_sheet = wb.worksheets[1]
    new_sheet = wb.create_sheet('FAK AMD 66 01-07-23')
    WorksheetCopy(source_worksheet=old_sheet, target_worksheet=new_sheet).copy_worksheet()
    offset = (-1) * (wb._sheets.index(new_sheet) - 1)
    wb.move_sheet(new_sheet, offset)
    new_sheet.sheet_view.zoomScale = 60
    for merge in list(new_sheet.merged_cells):
        range_string = str(merge)
        if range_string not in keep_merged:
            new_sheet.unmerge_cells(range_string)
    new_sheet.delete_rows(8, new_sheet.max_row-1)

    print("Inputting rate tables")
    rate_table_start = 8
    row_counter = rate_table_start
    for rate_table in rate_tables:
        rate_cell = new_sheet.cell(row=row_counter, column=1, value=rate_table.name)
        new_sheet.merge_cells(start_row=row_counter, start_column=columns['Group'], 
                              end_row=row_counter, end_column=columns['NAC'])
        row_counter += 1

        new_sheet.cell(row=row_counter, column=1, value='COMMODITY CODE')
        row_counter += 1

        code_counter = 0
        for code in rate_table.commodity_codes:
            code_cell = new_sheet.cell(row=row_counter+code_counter, column=columns['Group'], 
                                       value=code)
            styles.left_aligned_cell(code_cell)
            code_counter += 1
        
        for origin in rate_table.origins:
            origin_start = row_counter
            new_sheet.cell(row=origin_start, column=columns['Origin'], value=origin.name)
            for dest in origin.destinations:
                dest_cell = new_sheet.cell(row=row_counter, column=columns['Destination'], 
                                          value=dest.name)
                styles.centered_cell(dest_cell)
                via_cell = new_sheet.cell(row=row_counter, column=columns['Via'], 
                                          value=dest.via)
                styles.centered_cell(via_cell)
                term_cell = new_sheet.cell(row=row_counter, column=columns['Term'], 
                                          value=dest.term)
                styles.centered_cell(term_cell)
                for container_type, rate in dest.rates.items():
                    if(rate != 'nan'):
                        type_cell = new_sheet.cell(row=row_counter, column=columns['EQ'], 
                                       value=container_type)
                        styles.centered_cell(type_cell)
                        rate_cell = new_sheet.cell(row=row_counter, column=columns['O/F'], 
                                       value=rate)
                        styles.currency_cell(rate_cell)
                        formula = '=SUM(G' + str(row_counter) + ':Q' + str(row_counter) + ')'
                        subtotal_cell = new_sheet.cell(row=row_counter, column=columns['Sub Total'], 
                                       value=formula)
                        styles.currency_cell(subtotal_cell)
                        row_counter += 1
                row_counter += 1
            new_sheet.merge_cells(start_row=origin_start, start_column=columns['Origin'],
                                  end_row=row_counter-2, end_column=columns['Origin'])


    print("Saving file")
    wb.save(output_file)

if __name__ == "__main__":
    main()
